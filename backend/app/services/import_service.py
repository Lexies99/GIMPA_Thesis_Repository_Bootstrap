from __future__ import annotations

import csv
from io import BytesIO, StringIO
from pathlib import Path
import secrets
import string

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.student import Student
from app.schemas.student import StudentImportSummary
from app.schemas.user import UserUpdate
from app.services.email_service import send_notification_email
from app.services.user_service import (
    assign_role,
    create_user,
    get_user_by_email,
    get_user_by_school_id,
    is_allowed_institution_email,
    update_user,
)

VALID_STUDENT_BLOCK_CODES: dict[str, str] = {
    "A1": "SEMESTER 1 (SEPTEMBER INTAKE)",
    "A2": "SEMESTER 2 (SEPTEMBER INTAKE)",
    "A5": "SEMESTER 1 (FEBRUARY INTAKE)",
    "A6": "SEMESTER 2 (FEBRUARY INTAKE)",
    "B1": "QUARTER 1 (FEBRUARY INTAKE)",
    "B2": "QUARTER 2 (FEBRUARY INTAKE)",
    "B3": "QUARTER 3 (FEBRUARY INTAKE)",
    "B4": "QUARTER 4 (FEBRUARY INTAKE)",
    "BA": "QUARTER 1 REGULAR (SEPTEMBER INTAKE)",
    "BB": "QUARTER 2 REGULAR (SEPTEMBER INTAKE)",
    "BC": "QUARTER 3 REGULAR (SEPTEMBER INTAKE)",
    "BD": "QUARTER 4 REGULAR (SEPTEMBER INTAKE)",
    "F1": "SEMESTER 1 (FEBRUARY INTAKE)",
    "F2": "SEMESTER 2 (FEBRUARY INTAKE)",
    "JA": "MODULAR SESSION 1 (SEPTEMBER INTAKE)",
    "JB": "MODULAR SESSION 2 (SEPTEMBER INTAKE)",
    "JK": "MODULAR SESSION 3 FINAL (SEPTEMBER INTAKE)",
    "T1": "TRIMESTER 1 (FEBRUARY INTAKE)",
    "T2": "TRIMESTER 2 (FEBRUARY INTAKE)",
    "T3": "TRIMESTER 3 (FEBRUARY INTAKE)",
    "TA": "TRIMESTER 1 (SEPTEMBER INTAKE)",
    "TB": "TRIMESTER 2 (SEPTEMBER INTAKE)",
    "TC": "TRIMESTER 3 (SEPTEMBER INTAKE)",
    "X1": "SESSION 1 (FEBRUARY INTAKE)",
    "X2": "SESSION 2 (FEBRUARY INTAKE)",
}


def _normalize_key(value: str) -> str:
    return "".join(ch for ch in (value or "").strip().lower() if ch.isalnum())


def generate_default_password(length: int = 16) -> str:
    if length < 12:
        length = 12
    rng = secrets.SystemRandom()
    lowers = string.ascii_lowercase
    uppers = string.ascii_uppercase
    digits = string.digits
    specials = "!@#$%^&*()_-+=[]{}:;,.?"
    pool = lowers + uppers + digits + specials
    chars = [
        rng.choice(lowers),
        rng.choice(uppers),
        rng.choice(digits),
        rng.choice(specials),
    ]
    chars.extend(rng.choice(pool) for _ in range(length - len(chars)))
    rng.shuffle(chars)
    return "".join(chars)


def _map_value(row: dict[str, str], candidates: list[str]) -> str:
    normalized = {_normalize_key(k): v for k, v in row.items()}
    for candidate in candidates:
        value = normalized.get(_normalize_key(candidate), "").strip()
        if value:
            return value
    return ""


def _rows_from_csv_bytes(raw: bytes) -> list[dict[str, str]]:
    text = raw.decode("utf-8-sig", errors="ignore")
    reader = csv.DictReader(StringIO(text))
    return [{(k or "").strip(): (v or "").strip() for k, v in row.items()} for row in reader]


def _rows_from_xlsx_bytes(raw: bytes) -> list[dict[str, str]]:
    wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    raw_rows = list(ws.iter_rows(values_only=True))
    if not raw_rows:
        return []

    expected_headers = [
        "student name",
        "student id",
        "school email",
        "school",
        "department",
        "certification type",
        "block code",
        "year",
    ]
    lecturer_expected_headers = [
        "lecturer name",
        "lecturer id",
        "lecturer email",
        "adjunct email",
        "school",
        "department",
        "year",
    ]
    expected_keys = {_normalize_key(h) for h in expected_headers}

    def row_values(values: tuple) -> list[str]:
        return [str(cell or "").strip() for cell in values]

    # Try to locate a real header row.
    header_idx = 0
    found_header = False
    for idx, row in enumerate(raw_rows[:10]):
        vals = row_values(row)
        normalized = {_normalize_key(v) for v in vals if v}
        if len(normalized & expected_keys) >= 2:
            header_idx = idx
            found_header = True
            break

    items: list[dict[str, str]] = []

    # Support template files where row 1 is a banner and data starts directly.
    first_cell = str((raw_rows[0][0] if raw_rows[0] else "") or "").strip().lower()
    is_student_template_banner = "student bulk upload template" in first_cell
    is_lecturer_template_banner = "lecturer bulk upload template" in first_cell
    if not found_header and is_student_template_banner:
        for row in raw_rows[1:]:
            vals = row_values(row)
            if not any(vals):
                continue
            items.append({expected_headers[idx]: vals[idx] if idx < len(vals) else "" for idx in range(len(expected_headers))})
        return items
    if not found_header and is_lecturer_template_banner:
        for row in raw_rows[1:]:
            vals = row_values(row)
            if not any(vals):
                continue
            items.append(
                {
                    lecturer_expected_headers[idx]: vals[idx] if idx < len(vals) else ""
                    for idx in range(len(lecturer_expected_headers))
                }
            )
        return items

    headers = row_values(raw_rows[header_idx])
    for row in raw_rows[header_idx + 1:]:
        vals = row_values(row)
        if not any(vals):
            continue
        items.append({headers[idx]: vals[idx] if idx < len(vals) else "" for idx in range(len(headers))})
    return items


def load_rows_from_path(path: Path) -> list[dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _rows_from_csv_bytes(path.read_bytes())
    if suffix in {".xlsx", ".xlsm"}:
        return _rows_from_xlsx_bytes(path.read_bytes())
    raise ValueError(f"Unsupported file type: {suffix}")


def load_rows_from_upload(filename: str, raw: bytes) -> list[dict[str, str]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        return _rows_from_csv_bytes(raw)
    if suffix in {".xlsx", ".xlsm"}:
        return _rows_from_xlsx_bytes(raw)
    raise ValueError("Only .csv and .xlsx files are supported")


def _upsert_student(db: Session, row: dict[str, str]) -> tuple[bool, str, bool, bool]:
    student_id = _map_value(row, ["student id", "student_id", "id"])
    email = _map_value(row, ["school email", "email", "student email"])
    if not student_id:
        return False, "missing student id", False, False
    if not email:
        return False, f"{student_id}: missing school email", False, False
    if not is_allowed_institution_email(email):
        return False, f"{student_id}: invalid institution email", False, False

    certification_type = _map_value(
        row,
        ["certification type", "certification", "programme type", "program type", "award type", "degree type"],
    )
    block_code_raw = _map_value(row, ["block code", "block", "session code", "intake block"])
    block_code = block_code_raw.strip().upper() if block_code_raw else ""
    if block_code and block_code not in VALID_STUDENT_BLOCK_CODES:
        return False, f"{student_id}: invalid block code '{block_code}'", False, False

    year_raw = _map_value(row, ["year"])
    year: int | None = None
    if year_raw:
        try:
            year = int(year_raw)
        except ValueError:
            return False, f"{student_id}: invalid year '{year_raw}'", False, False

    existing = db.query(Student).filter(Student.student_id == student_id).first()
    payload = {
        "student_id": student_id,
        "full_name": _map_value(row, ["student name", "name", "full name"]),
        "email": email.lower(),
        "school": _map_value(row, ["school"]),
        "department": _map_value(row, ["department"]),
        "program": _map_value(row, ["program", "programme", "course", "major", "discipline"]),
        "certification_type": certification_type or None,
        "block_code": block_code or None,
        "year": year,
    }
    if not payload["full_name"]:
        return False, f"{student_id}: missing student name", False, False
    if existing:
        for key, value in payload.items():
            setattr(existing, key, value)
    else:
        db.add(Student(**payload))

    existing_user = get_user_by_email(db, email.lower())
    if existing_user:
        if existing_user.school_id and existing_user.school_id != student_id:
            return False, f"{student_id}: email already mapped to another school id", False, False
        updated_user = update_user(
            db,
            existing_user,
            UserUpdate(
                full_name=payload["full_name"],
                school_id=student_id,
                school=payload["school"] or existing_user.school,
                department=payload["department"] or existing_user.department,
                program=payload["program"] or existing_user.program,
                is_active=True,
            ),
        )
        assign_role(db, updated_user, "student")
        sent = send_notification_email(
            to_email=updated_user.email,
            to_name=updated_user.full_name,
            subject="GIMPA Thesis Management System Account Update Confirmation",
            message=(
                "Your student account details have been updated successfully in GIMPA Thesis Management System.\n\n"
                "Updated details:\n"
                f"- Email: {updated_user.email}\n"
                f"- School ID: {updated_user.school_id or student_id}\n\n"
                "If any of these details are incorrect, please contact your administrator."
            ),
        )
        if not sent:
            return True, f"{student_id}: student updated but email notification failed", False, True
        return True, student_id, True, False

    existing_by_school_id = get_user_by_school_id(db, student_id)
    if existing_by_school_id:
        return False, f"{student_id}: school id already used by another account", False, False

    generated_password = generate_default_password()
    user = create_user(
        db=db,
        email=email.lower(),
        role="student",
        school_id=student_id,
        school=payload["school"] or None,
        password=generated_password,
        full_name=payload["full_name"] or None,
        department=payload["department"] or None,
        must_change_password=True,
    )
    user.program = payload["program"] or None
    user.is_active = True
    db.add(user)
    db.commit()
    db.refresh(user)

    sent = send_notification_email(
        to_email=user.email,
        to_name=user.full_name,
        subject="Welcome to GIMPA Thesis Management System - Student Account Created",
        message=(
            "Your student account has been created successfully in GIMPA Thesis Management System.\n\n"
            "Account details:\n"
            f"- Email: {user.email}\n"
            f"- School ID: {user.school_id or student_id}\n"
            f"- Temporary Password: {generated_password}\n\n"
            "Next steps:\n"
            "1. Sign in to your account.\n"
            "2. Change your temporary password immediately.\n"
            "3. Complete your profile if needed."
        ),
    )
    if not sent:
        return True, f"{student_id}: account created but email could not be delivered", False, True
    return True, student_id, True, False


def _upsert_staff_user(db: Session, row: dict[str, str], default_role: str) -> tuple[bool, str, bool, bool]:
    name = _map_value(row, ["lecturer name", "name", "full name"])
    school_id = _map_value(row, ["lecturer id", "lecture id", "staff id", "school id", "id"])
    email = _map_value(row, ["lecturer email", "lecture email", "school email", "email"])
    if not email:
        return False, "missing email", False, False
    if not is_allowed_institution_email(email):
        return False, f"{email}: invalid institution email", False, False

    role_value = _map_value(row, ["role", "roles"]) or default_role
    roles = [item.strip().lower() for item in role_value.split("|") if item.strip()]
    if not roles:
        roles = [default_role]

    existing = get_user_by_email(db, email)
    if existing:
        updated = update_user(
            db,
            existing,
            UserUpdate(
                full_name=name or existing.full_name,
                school_id=school_id or existing.school_id,
                school=_map_value(row, ["school"]) or existing.school,
                department=_map_value(row, ["department"]) or existing.department,
                is_active=True,
            ),
        )
        for role in roles:
            assign_role(db, updated, role)
        return True, email, False, False

    if school_id and get_user_by_school_id(db, school_id):
        return False, f"{email}: school_id already in use", False, False

    generated_password = generate_default_password()

    user = create_user(
        db=db,
        email=email,
        role=roles[0],
        school_id=school_id or None,
        school=_map_value(row, ["school"]) or None,
        password=generated_password,
        full_name=name or None,
        department=_map_value(row, ["department"]) or None,
        must_change_password=True,
    )
    user.is_active = True
    db.add(user)
    db.commit()
    db.refresh(user)
    for role in roles[1:]:
        assign_role(db, user, role)

    sent = send_notification_email(
        to_email=user.email,
        to_name=user.full_name,
        subject="Welcome to GIMPA Thesis Management System - Account Created",
        message=(
            "Your account has been created successfully in GIMPA Thesis Management System.\n\n"
            "Account details:\n"
            f"- Email: {user.email}\n"
            f"- Temporary Password: {generated_password}\n\n"
            "Next steps:\n"
            "1. Sign in to your account.\n"
            "2. Change your temporary password immediately.\n"
            "3. Review your assigned role and profile information."
        ),
    )
    if not sent:
        return True, f"{email}: account created but email could not be delivered (check SMTP settings)", False, True
    return True, email, True, False


def import_students(db: Session, rows: list[dict[str, str]]) -> StudentImportSummary:
    ok = 0
    emailed_sent = 0
    emailed_failed = 0
    errors: list[str] = []
    for row in rows:
        success, info, mail_sent, mail_failed = _upsert_student(db, row)
        if success:
            ok += 1
            if mail_sent:
                emailed_sent += 1
            if mail_failed:
                emailed_failed += 1
                errors.append(info)
        else:
            errors.append(info)
    db.commit()
    return StudentImportSummary(
        imported_or_updated=ok,
        emailed_sent=emailed_sent,
        emailed_failed=emailed_failed,
        errors=errors,
    )


def import_staff_accounts(
    db: Session,
    rows: list[dict[str, str]],
    *,
    default_role: str,
) -> StudentImportSummary:
    ok = 0
    emailed_sent = 0
    emailed_failed = 0
    errors: list[str] = []
    for row in rows:
        success, info, mail_sent, mail_failed = _upsert_staff_user(db, row, default_role=default_role)
        if success:
            ok += 1
            if mail_sent:
                emailed_sent += 1
            if mail_failed:
                emailed_failed += 1
                errors.append(info)
        else:
            errors.append(info)
    db.commit()
    return StudentImportSummary(
        imported_or_updated=ok,
        emailed_sent=emailed_sent,
        emailed_failed=emailed_failed,
        errors=errors,
    )
